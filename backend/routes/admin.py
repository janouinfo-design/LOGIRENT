from fastapi import APIRouter, HTTPException, Depends, File, UploadFile, Request
from typing import Optional
from datetime import datetime, timedelta
import uuid
import asyncio
import base64
import logging
import json
import os

from database import db
from models import AdminStats, AdminUserUpdate, Base64UserPhoto, PaymentTransaction, PricingTiersUpdate
from pydantic import BaseModel as PydanticBaseModel
from deps import get_admin_user, get_agency_admin, get_current_user, hash_password
from utils.notifications import create_notification
from utils.helpers import verify_document_with_ai
from utils.helpers import extract_document_ocr
from utils.email import (
    send_email, send_reservation_confirmation, send_payment_confirmation,
    generate_status_change_email
)

logger = logging.getLogger(__name__)
router = APIRouter()


@router.get("/admin/stats", response_model=AdminStats)
async def get_admin_stats(user: dict = Depends(get_admin_user)):
    agency_id = user.get('agency_id')
    is_super = user.get('role') == 'super_admin'

    vf = {} if is_super else {"agency_id": agency_id}
    rf = {} if is_super else {"agency_id": agency_id}

    total_vehicles = await db.vehicles.count_documents(vf)
    total_reservations = await db.reservations.count_documents(rf)

    if is_super:
        total_users = await db.users.count_documents({})
        total_payments = await db.payment_transactions.count_documents({"payment_status": "paid"})
    else:
        user_ids = await db.reservations.distinct("user_id", rf)
        total_users = len(user_ids)
        total_payments = await db.reservations.count_documents({**rf, "payment_status": "paid"})

    revenue_match = {"payment_status": "paid"}
    if not is_super:
        revenue_match["agency_id"] = agency_id
    pipeline = [{"$match": revenue_match}, {"$group": {"_id": None, "total": {"$sum": "$total_price"}}}]
    revenue_result = await db.reservations.aggregate(pipeline).to_list(1)
    total_revenue = revenue_result[0]["total"] if revenue_result else 0

    status_pipeline = [{"$match": rf}, {"$group": {"_id": "$status", "count": {"$sum": 1}}}]
    status_result = await db.reservations.aggregate(status_pipeline).to_list(10)
    reservations_by_status = {item["_id"]: item["count"] for item in status_result}

    top_pipeline = [{"$match": rf}, {"$group": {"_id": "$vehicle_id", "rental_count": {"$sum": 1}}}, {"$sort": {"rental_count": -1}}, {"$limit": 5}]
    top_result = await db.reservations.aggregate(top_pipeline).to_list(5)
    top_vehicles = []
    for item in top_result:
        vehicle = await db.vehicles.find_one({"id": item["_id"]})
        if vehicle:
            top_vehicles.append({"id": vehicle["id"], "name": f"{vehicle['brand']} {vehicle['model']}", "rental_count": item["rental_count"]})

    six_months_ago = datetime.utcnow() - timedelta(days=180)
    monthly_match = {"payment_status": "paid", "created_at": {"$gte": six_months_ago}}
    if not is_super:
        monthly_match["agency_id"] = agency_id
    monthly_pipeline = [{"$match": monthly_match}, {"$group": {"_id": {"year": {"$year": "$created_at"}, "month": {"$month": "$created_at"}}, "revenue": {"$sum": "$total_price"}, "count": {"$sum": 1}}}, {"$sort": {"_id.year": 1, "_id.month": 1}}]
    monthly_result = await db.reservations.aggregate(monthly_pipeline).to_list(12)
    revenue_by_month = [{"month": datetime(item["_id"]["year"], item["_id"]["month"], 1).strftime("%b %Y"), "revenue": item["revenue"], "reservations": item["count"]} for item in monthly_result]

    return AdminStats(
        total_vehicles=total_vehicles, total_users=total_users,
        total_reservations=total_reservations, total_payments=total_payments,
        total_revenue=total_revenue, reservations_by_status=reservations_by_status,
        top_vehicles=top_vehicles, revenue_by_month=revenue_by_month
    )


@router.get("/admin/stats/advanced")
async def get_advanced_stats(user: dict = Depends(get_admin_user)):
    agency_id = user.get('agency_id')
    is_super = user.get('role') == 'super_admin'
    rf = {} if is_super else {"agency_id": agency_id}
    vf = {} if is_super else {"agency_id": agency_id}

    now = datetime.utcnow()
    thirty_days_ago = now - timedelta(days=30)
    this_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_month_start = (this_month_start - timedelta(days=1)).replace(day=1)

    rev_this = await db.reservations.aggregate([{"$match": {**rf, "payment_status": "paid", "created_at": {"$gte": this_month_start}}}, {"$group": {"_id": None, "total": {"$sum": "$total_price"}, "count": {"$sum": 1}}}]).to_list(1)
    rev_last = await db.reservations.aggregate([{"$match": {**rf, "payment_status": "paid", "created_at": {"$gte": last_month_start, "$lt": this_month_start}}}, {"$group": {"_id": None, "total": {"$sum": "$total_price"}, "count": {"$sum": 1}}}]).to_list(1)
    revenue_this_month = rev_this[0]["total"] if rev_this else 0
    revenue_last_month = rev_last[0]["total"] if rev_last else 0
    reservations_this_month = rev_this[0]["count"] if rev_this else 0
    reservations_last_month = rev_last[0]["count"] if rev_last else 0

    avg_dur_res = await db.reservations.aggregate([{"$match": {**rf, "status": {"$in": ["confirmed", "active", "completed"]}}}, {"$group": {"_id": None, "avg_days": {"$avg": "$total_days"}}}]).to_list(1)
    avg_booking_duration = round(avg_dur_res[0]["avg_days"], 1) if avg_dur_res else 0

    avg_rev_res = await db.reservations.aggregate([{"$match": {**rf, "payment_status": "paid"}}, {"$group": {"_id": None, "avg_rev": {"$avg": "$total_price"}}}]).to_list(1)
    avg_revenue_per_reservation = round(avg_rev_res[0]["avg_rev"], 2) if avg_rev_res else 0

    vehicles = await db.vehicles.find(vf, {"_id": 0, "id": 1, "brand": 1, "model": 1}).to_list(200)
    vehicle_utilization = []
    for v in vehicles[:20]:
        res_list = await db.reservations.find({"vehicle_id": v["id"], "status": {"$in": ["confirmed", "active", "completed"]}, "start_date": {"$lte": now}, "end_date": {"$gte": thirty_days_ago}}, {"_id": 0, "start_date": 1, "end_date": 1}).to_list(100)
        booked_days = 0
        for r in res_list:
            s = max(r["start_date"], thirty_days_ago) if isinstance(r["start_date"], datetime) else thirty_days_ago
            e = min(r["end_date"], now) if isinstance(r["end_date"], datetime) else now
            booked_days += max(0, (e - s).days)
        rate = min(100, round((booked_days / 30) * 100))
        vehicle_utilization.append({"id": v["id"], "name": f"{v['brand']} {v['model']}", "utilization": rate, "booked_days": booked_days})
    vehicle_utilization.sort(key=lambda x: x["utilization"], reverse=True)

    rev_per_vehicle_result = await db.reservations.aggregate([{"$match": {**rf, "payment_status": "paid"}}, {"$group": {"_id": "$vehicle_id", "revenue": {"$sum": "$total_price"}, "count": {"$sum": 1}}}, {"$sort": {"revenue": -1}}, {"$limit": 10}]).to_list(10)
    revenue_per_vehicle = []
    for item in rev_per_vehicle_result:
        veh = await db.vehicles.find_one({"id": item["_id"]})
        if veh:
            revenue_per_vehicle.append({"id": veh["id"], "name": f"{veh['brand']} {veh['model']}", "revenue": item["revenue"], "bookings": item["count"]})

    daily_result = await db.reservations.aggregate([{"$match": {**rf, "payment_status": "paid", "created_at": {"$gte": thirty_days_ago}}}, {"$group": {"_id": {"$dateToString": {"format": "%Y-%m-%d", "date": "$created_at"}}, "revenue": {"$sum": "$total_price"}, "count": {"$sum": 1}}}, {"$sort": {"_id": 1}}]).to_list(31)
    daily_revenue = [{"date": item["_id"], "revenue": item["revenue"], "bookings": item["count"]} for item in daily_result]

    new_clients_30d = await db.users.count_documents({"role": "client", "created_at": {"$gte": thirty_days_ago}}) if is_super else 0
    if not is_super:
        new_user_ids = await db.reservations.distinct("user_id", {**rf, "created_at": {"$gte": thirty_days_ago}})
        new_clients_30d = len(new_user_ids)

    pm_result = await db.reservations.aggregate([{"$match": {**rf, "payment_status": "paid"}}, {"$group": {"_id": {"$ifNull": ["$payment_method", "card"]}, "count": {"$sum": 1}, "total": {"$sum": "$total_price"}}}]).to_list(10)
    payment_methods = [{"method": item["_id"] or "card", "count": item["count"], "total": item["total"]} for item in pm_result]

    total_res = await db.reservations.count_documents(rf) or 1
    cancelled_res = await db.reservations.count_documents({**rf, "status": "cancelled"})
    cancellation_rate = round((cancelled_res / total_res) * 100, 1)

    eight_weeks_ago = now - timedelta(weeks=8)
    weekly_result = await db.reservations.aggregate([{"$match": {**rf, "created_at": {"$gte": eight_weeks_ago}}}, {"$group": {"_id": {"$isoWeek": "$created_at"}, "count": {"$sum": 1}, "revenue": {"$sum": {"$cond": [{"$eq": ["$payment_status", "paid"]}, "$total_price", 0]}}}}, {"$sort": {"_id": 1}}]).to_list(8)
    weekly_trends = [{"week": item["_id"], "bookings": item["count"], "revenue": item["revenue"]} for item in weekly_result]

    return {
        "revenue_this_month": revenue_this_month, "revenue_last_month": revenue_last_month,
        "revenue_change_pct": round(((revenue_this_month - revenue_last_month) / revenue_last_month * 100) if revenue_last_month > 0 else 0, 1),
        "reservations_this_month": reservations_this_month, "reservations_last_month": reservations_last_month,
        "avg_booking_duration": avg_booking_duration, "avg_revenue_per_reservation": avg_revenue_per_reservation,
        "vehicle_utilization": vehicle_utilization, "revenue_per_vehicle": revenue_per_vehicle,
        "daily_revenue": daily_revenue, "new_clients_30d": new_clients_30d,
        "payment_methods": payment_methods, "cancellation_rate": cancellation_rate,
        "weekly_trends": weekly_trends,
    }


@router.get("/admin/stats/top-clients")
async def get_top_clients(user: dict = Depends(get_admin_user)):
    agency_id = user.get('agency_id')
    is_super = user.get('role') == 'super_admin'
    rf = {} if is_super else {"agency_id": agency_id}

    pipeline = [
        {"$match": {**rf, "payment_status": "paid"}},
        {"$group": {"_id": "$user_id", "total_spent": {"$sum": "$total_price"}, "bookings": {"$sum": 1}}},
        {"$sort": {"total_spent": -1}},
        {"$limit": 10},
    ]
    top = await db.reservations.aggregate(pipeline).to_list(10)
    result = []
    for item in top:
        u = await db.users.find_one({"id": item["_id"]}, {"_id": 0, "id": 1, "name": 1, "email": 1, "client_rating": 1})
        if u:
            result.append({"id": u["id"], "name": u.get("name", ""), "email": u.get("email", ""),
                           "rating": u.get("client_rating", "neutral"),
                           "total_spent": item["total_spent"], "bookings": item["bookings"]})
    return result


@router.get("/admin/stats/agency-comparison")
async def get_agency_comparison(user: dict = Depends(get_admin_user)):
    if user.get('role') != 'super_admin':
        raise HTTPException(status_code=403, detail="Super admin only")

    agencies = await db.agencies.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(50)
    result = []
    for a in agencies:
        rev = await db.reservations.aggregate([
            {"$match": {"agency_id": a["id"], "payment_status": "paid"}},
            {"$group": {"_id": None, "revenue": {"$sum": "$total_price"}, "bookings": {"$sum": 1}}}
        ]).to_list(1)
        vehicle_count = await db.vehicles.count_documents({"agency_id": a["id"]})
        result.append({
            "id": a["id"], "name": a["name"],
            "revenue": rev[0]["revenue"] if rev else 0,
            "bookings": rev[0]["bookings"] if rev else 0,
            "vehicles": vehicle_count,
        })
    result.sort(key=lambda x: x["revenue"], reverse=True)
    return result


@router.get("/admin/stats/tier-analytics")
async def get_tier_analytics(user: dict = Depends(get_admin_user)):
    agency_id = user.get('agency_id')
    is_super = user.get('role') == 'super_admin'
    rf = {} if is_super else {"agency_id": agency_id}

    pipeline = [
        {"$match": {**rf, "selected_tier": {"$ne": None}}},
        {"$group": {
            "_id": "$selected_tier.name",
            "bookings": {"$sum": 1},
            "revenue": {"$sum": "$total_price"},
            "avg_price": {"$avg": "$total_price"},
        }},
        {"$sort": {"revenue": -1}},
    ]
    tier_stats = await db.reservations.aggregate(pipeline).to_list(20)

    total_with_tier = await db.reservations.count_documents({**rf, "selected_tier": {"$ne": None}})
    total_without_tier = await db.reservations.count_documents({**rf, "selected_tier": None})

    return {
        "tier_stats": [{"name": t["_id"] or "Sans forfait", "bookings": t["bookings"], "revenue": t["revenue"], "avg_price": round(t["avg_price"], 2)} for t in tier_stats],
        "with_tier": total_with_tier,
        "without_tier": total_without_tier,
    }


@router.get("/admin/stats/revenue-forecast")
async def get_revenue_forecast(user: dict = Depends(get_admin_user)):
    agency_id = user.get('agency_id')
    is_super = user.get('role') == 'super_admin'
    rf = {} if is_super else {"agency_id": agency_id}

    now = datetime.utcnow()
    twelve_months_ago = now - timedelta(days=365)

    monthly_pipeline = [
        {"$match": {**rf, "payment_status": "paid", "created_at": {"$gte": twelve_months_ago}}},
        {"$group": {
            "_id": {"year": {"$year": "$created_at"}, "month": {"$month": "$created_at"}},
            "revenue": {"$sum": "$total_price"},
            "bookings": {"$sum": 1}
        }},
        {"$sort": {"_id.year": 1, "_id.month": 1}}
    ]
    monthly_result = await db.reservations.aggregate(monthly_pipeline).to_list(12)

    historical = []
    for item in monthly_result:
        month_str = datetime(item["_id"]["year"], item["_id"]["month"], 1).strftime("%Y-%m")
        historical.append({
            "month": month_str,
            "revenue": round(item["revenue"], 2),
            "bookings": item["bookings"]
        })

    total_vehicles = await db.vehicles.count_documents({} if is_super else {"agency_id": agency_id})
    avg_price_res = await db.vehicles.aggregate([
        {"$match": {} if is_super else {"agency_id": agency_id}},
        {"$group": {"_id": None, "avg_price": {"$avg": "$price_per_day"}}}
    ]).to_list(1)
    avg_daily_price = round(avg_price_res[0]["avg_price"], 2) if avg_price_res else 0

    forecast = []
    analysis = ""
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        api_key = os.environ.get("EMERGENT_LLM_KEY")
        if api_key and historical:
            chat = LlmChat(
                api_key=api_key,
                session_id=f"forecast-{agency_id or 'global'}-{now.strftime('%Y%m%d%H')}",
                system_message="Tu es un analyste financier expert en location de véhicules. Réponds en français. Analyse les données historiques et génère des prévisions précises."
            ).with_model("openai", "gpt-5.2")

            prompt = f"""Voici les données historiques de revenus mensuels d'une agence de location de véhicules:

{json.dumps(historical, ensure_ascii=False)}

Contexte:
- Nombre de véhicules: {total_vehicles}
- Prix moyen journalier: CHF {avg_daily_price}
- Mois actuel: {now.strftime('%Y-%m')}

Génère une prévision pour les 3 prochains mois et une analyse courte.
Réponds UNIQUEMENT en JSON valide avec ce format exact:
{{
  "forecast": [
    {{"month": "YYYY-MM", "revenue": number, "bookings": number, "confidence": number_between_0_and_1}}
  ],
  "analysis": "string courte en français (2-3 phrases max)",
  "trend": "up" ou "down" ou "stable"
}}"""

            user_msg = UserMessage(text=prompt)
            response = await chat.send_message(user_msg)

            try:
                cleaned = response.strip()
                if cleaned.startswith("```"):
                    cleaned = cleaned.split("\n", 1)[1] if "\n" in cleaned else cleaned[3:]
                    if cleaned.endswith("```"):
                        cleaned = cleaned[:-3]
                    cleaned = cleaned.strip()
                    if cleaned.startswith("json"):
                        cleaned = cleaned[4:].strip()
                parsed = json.loads(cleaned)
                forecast = parsed.get("forecast", [])
                analysis = parsed.get("analysis", "")
                trend = parsed.get("trend", "stable")
            except (json.JSONDecodeError, AttributeError):
                trend = "stable"
                analysis = "Prévision non disponible - données insuffisantes."
        else:
            trend = "stable"
            if not historical:
                analysis = "Aucune donnée historique disponible pour générer une prévision."
            else:
                analysis = "Clé API non configurée."
    except Exception as e:
        logger.error(f"Revenue forecast AI error: {e}")
        trend = "stable"
        analysis = "Prévision temporairement indisponible."

    return {
        "historical": historical,
        "forecast": forecast,
        "analysis": analysis,
        "trend": trend,
        "total_vehicles": total_vehicles,
        "avg_daily_price": avg_daily_price
    }



@router.get("/admin/users")
async def get_admin_users(skip: int = 0, limit: int = 20, user: dict = Depends(get_admin_user)):
    agency_id = user.get('agency_id')
    is_super = user.get('role') == 'super_admin'

    if is_super:
        users = await db.users.find({}, {"password_hash": 0}).skip(skip).limit(limit).to_list(limit)
        total = await db.users.count_documents({})
    else:
        user_ids_from_reservations = await db.reservations.distinct("user_id", {"agency_id": agency_id})
        query = {"$or": [{"id": {"$in": user_ids_from_reservations}}, {"agency_id": agency_id, "role": "client"}]}
        users = await db.users.find(query, {"password_hash": 0}).skip(skip).limit(limit).to_list(limit)
        total = await db.users.count_documents(query)

    for u in users:
        u['_id'] = str(u['_id'])

    # Batch fetch reservation counts
    user_ids = [u['id'] for u in users]
    pipeline = [
        {"$match": {"user_id": {"$in": user_ids}} if is_super else {"user_id": {"$in": user_ids}, "agency_id": agency_id}},
        {"$group": {"_id": "$user_id", "count": {"$sum": 1}}}
    ]
    counts = {r['_id']: r['count'] async for r in db.reservations.aggregate(pipeline)}
    for u in users:
        u['reservation_count'] = counts.get(u['id'], 0)

    return {"users": users, "total": total}


@router.put("/admin/users/{user_id}/block")
async def block_user(user_id: str, user: dict = Depends(get_admin_user)):
    target_user = await db.users.find_one({"id": user_id})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    new_status = not target_user.get('blocked', False)
    await db.users.update_one({"id": user_id}, {"$set": {"blocked": new_status}})
    return {"message": f"User {'blocked' if new_status else 'unblocked'}"}


@router.put("/admin/users/{user_id}")
async def update_user_admin(user_id: str, update_data: AdminUserUpdate, user: dict = Depends(get_admin_user)):
    target_user = await db.users.find_one({"id": user_id})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")

    update_dict = {k: v for k, v in update_data.dict().items() if v is not None}
    if update_dict:
        await db.users.update_one({"id": user_id}, {"$set": update_dict})

    updated_user = await db.users.find_one({"id": user_id}, {"password_hash": 0})
    updated_user['_id'] = str(updated_user['_id'])
    return {"message": "User updated successfully", "user": updated_user}


@router.delete("/admin/users/{user_id}")
async def delete_user_admin(user_id: str, force: bool = False, user: dict = Depends(get_admin_user)):
    target = await db.users.find_one({"id": user_id}, {"_id": 0, "role": 1, "name": 1})
    if not target:
        raise HTTPException(status_code=404, detail="Client non trouve")
    if target.get('role') in ('admin', 'super_admin'):
        raise HTTPException(status_code=403, detail="Impossible de supprimer un admin")
    # Check active reservations
    active = await db.reservations.count_documents({"user_id": user_id, "status": {"$in": ["active", "confirmed", "pending"]}})
    if active > 0 and not force:
        raise HTTPException(status_code=400, detail=f"Ce client a {active} reservation(s) active(s). Utilisez force=true pour supprimer et annuler ses reservations.")
    if active > 0 and force:
        await db.reservations.update_many(
            {"user_id": user_id, "status": {"$in": ["active", "confirmed", "pending"]}},
            {"$set": {"status": "cancelled"}}
        )
    await db.users.delete_one({"id": user_id})
    return {"message": f"Client {target.get('name','')} supprime"}



class AdminDocUpload(PydanticBaseModel):
    image_data: str
    doc_type: str  # id, id_back, license, license_back


@router.post("/admin/client/{client_id}/document")
async def admin_upload_client_document(client_id: str, body: AdminDocUpload, user: dict = Depends(get_admin_user)):
    target = await db.users.find_one({"id": client_id})
    if not target:
        raise HTTPException(status_code=404, detail="Client not found")

    doc_type_map = {"id": "id", "id_back": "id", "license": "license", "license_back": "license"}
    ai_type = doc_type_map.get(body.doc_type, "id")
    verification = await verify_document_with_ai(body.image_data, ai_type)

    if not verification.get("is_valid", True) and verification.get("confidence", 0) > 60:
        return {"message": "Document rejeté", "verification": verification, "rejected": True}

    field_map = {"id": "id_photo", "id_back": "id_photo_back", "license": "license_photo", "license_back": "license_photo_back"}
    field = field_map.get(body.doc_type, "id_photo")
    verif_field = "id_verification" if "id" in body.doc_type else "license_verification"

    update = {"$set": {field: body.image_data}}
    if body.doc_type in ("id", "license"):
        update["$set"][verif_field] = {
            "is_valid": verification.get("is_valid", True),
            "confidence": verification.get("confidence", 0),
            "reason": verification.get("reason", ""),
            "face": verification.get("face", "recto"),
            "is_blurry": verification.get("is_blurry", False),
            "is_expired": verification.get("is_expired", False),
            "quality_score": verification.get("quality_score", 0),
            "warnings": verification.get("warnings", []),
            "rejection_reasons": verification.get("rejection_reasons", []),
            "verified_at": datetime.utcnow().isoformat(),
        }

    await db.users.update_one({"id": client_id}, update)
    return {"message": "Document uploadé", "verification": verification}


@router.post("/admin/vehicle-inspection/ai-zone")
async def ai_inspect_zone(request: Request, user: dict = Depends(get_admin_user)):
    """AI analyzes a single zone photo for damages"""
    body = await request.json()
    image_data = body.get('image_data', '')
    zone_name = body.get('zone_name', 'zone')
    if not image_data:
        raise HTTPException(status_code=400, detail="Image requise")
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        api_key = os.environ.get("EMERGENT_LLM_KEY")
        if not api_key:
            return {"damages": [], "description": "IA non configuree", "severity": "unknown"}
        chat = LlmChat(
            api_key=api_key,
            session_id=f"inspect-zone-{uuid.uuid4().hex[:8]}",
            system_message="Tu es un expert en inspection de vehicules. Tu analyses des photos de zones de vehicules pour detecter les dommages. Reponds UNIQUEMENT en JSON valide."
        ).with_model("openai", "gpt-5.2")
        prompt = f"""Analyse cette photo de la zone "{zone_name}" d'un vehicule de location.
Detecte tous les dommages visibles (rayures, bosses, eclats, fissures, usure, etc.).
Reponds UNIQUEMENT en JSON:
{{
  "has_damage": true/false,
  "description": "description courte en francais des dommages detectes",
  "damages": ["rayure 5cm", "bosse legere"],
  "severity": "none" | "light" | "medium" | "severe",
  "confidence": 0-100
}}"""
        user_msg = UserMessage(text=prompt, image_url=image_data)
        response = await chat.send_message(user_msg)
        result = json.loads(response.text.strip().replace("```json", "").replace("```", ""))
        return result
    except json.JSONDecodeError:
        return {"has_damage": False, "description": "Analyse non concluante", "severity": "unknown", "confidence": 0}
    except Exception as e:
        logger.error(f"AI zone inspection error: {e}")
        return {"has_damage": False, "description": f"Erreur IA: {str(e)[:100]}", "severity": "unknown", "confidence": 0}


@router.post("/admin/vehicle-inspection/ai-global")
async def ai_inspect_global(request: Request, user: dict = Depends(get_admin_user)):
    """AI analyzes a global vehicle photo to detect all damaged zones"""
    body = await request.json()
    image_data = body.get('image_data', '')
    if not image_data:
        raise HTTPException(status_code=400, detail="Image requise")
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        api_key = os.environ.get("EMERGENT_LLM_KEY")
        if not api_key:
            return {"zones": [], "summary": "IA non configuree"}
        chat = LlmChat(
            api_key=api_key,
            session_id=f"inspect-global-{uuid.uuid4().hex[:8]}",
            system_message="Tu es un expert en inspection de vehicules. Tu analyses des photos de vehicules pour detecter les dommages sur chaque zone. Reponds UNIQUEMENT en JSON valide."
        ).with_model("openai", "gpt-5.2")
        zones_list = "pare_chocs_avant, ailiere_gauche_avant, toit, ailiere_droit_avant, porte_avant_gauche, porte_avant_droite, porte_arriere_gauche, coffre, porte_arriere_droite, ailiere_gauche_arriere, pare_chocs_arriere, ailier_droit_arriere"
        prompt = f"""Analyse cette photo globale d'un vehicule de location.
Identifie TOUTES les zones endommagees parmi: {zones_list}
Pour chaque zone avec dommage, decris le dommage et sa severite.
Reponds UNIQUEMENT en JSON:
{{
  "zones": [
    {{"zone_key": "pare_chocs_avant", "description": "Rayure profonde 15cm", "severity": "medium"}},
  ],
  "summary": "Resume global en francais",
  "overall_condition": "excellent" | "good" | "fair" | "poor",
  "confidence": 0-100
}}
Si aucun dommage: {{"zones": [], "summary": "Aucun dommage detecte", "overall_condition": "excellent", "confidence": 85}}"""
        user_msg = UserMessage(text=prompt, image_url=image_data)
        response = await chat.send_message(user_msg)
        result = json.loads(response.text.strip().replace("```json", "").replace("```", ""))
        return result
    except json.JSONDecodeError:
        return {"zones": [], "summary": "Analyse non concluante", "overall_condition": "unknown", "confidence": 0}
    except Exception as e:
        logger.error(f"AI global inspection error: {e}")
        return {"zones": [], "summary": f"Erreur IA: {str(e)[:100]}", "overall_condition": "unknown", "confidence": 0}


@router.put("/admin/users/{user_id}/rating")
async def update_user_rating(user_id: str, rating: str, user: dict = Depends(get_admin_user)):
    valid_ratings = ["good", "bad", "neutral", "vip", "blocked"]
    if rating not in valid_ratings:
        raise HTTPException(status_code=400, detail=f"Invalid rating. Must be one of: {valid_ratings}")
    target_user = await db.users.find_one({"id": user_id})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    await db.users.update_one({"id": user_id}, {"$set": {"client_rating": rating}})
    return {"message": f"User rating updated to {rating}"}


@router.post("/admin/users/{user_id}/photo")
async def upload_user_photo_admin(user_id: str, data: Base64UserPhoto, user: dict = Depends(get_admin_user)):
    target_user = await db.users.find_one({"id": user_id})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    data_uri = f"data:{data.content_type};base64,{data.image}"
    await db.users.update_one({"id": user_id}, {"$set": {"profile_photo": data_uri}})
    return {"message": "Photo uploaded successfully", "photo": data_uri}


@router.post("/admin/users/{user_id}/id-photo")
async def upload_user_id_photo_admin(user_id: str, data: Base64UserPhoto, user: dict = Depends(get_admin_user)):
    target_user = await db.users.find_one({"id": user_id})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    data_uri = f"data:{data.content_type};base64,{data.image}"
    await db.users.update_one({"id": user_id}, {"$set": {"id_photo": data_uri}})
    return {"message": "ID photo uploaded successfully", "photo": data_uri}


@router.post("/admin/users/{user_id}/license-photo")
async def upload_user_license_photo_admin(user_id: str, data: Base64UserPhoto, user: dict = Depends(get_admin_user)):
    target_user = await db.users.find_one({"id": user_id})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    data_uri = f"data:{data.content_type};base64,{data.image}"
    await db.users.update_one({"id": user_id}, {"$set": {"license_photo": data_uri}})
    return {"message": "License photo uploaded successfully", "photo": data_uri}


@router.get("/admin/users/{user_id}")
async def get_user_details_admin(user_id: str, user: dict = Depends(get_admin_user)):
    target_user = await db.users.find_one({"id": user_id}, {"password_hash": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")

    reservations = await db.reservations.find({"user_id": user_id}).to_list(100)
    total_spent = sum(r.get('total_price', 0) for r in reservations if r.get('payment_status') == 'paid')

    target_user['_id'] = str(target_user['_id'])
    target_user['total_spent'] = total_spent
    target_user['total_reservations'] = len(reservations)
    for r in reservations:
        r['_id'] = str(r['_id'])
    target_user['reservations'] = reservations
    return target_user


@router.post("/admin/import-users")
async def import_users_from_excel(file: UploadFile = File(...), user: dict = Depends(get_admin_user)):
    import io
    import zipfile

    filename = file.filename or ""
    content = await file.read()

    photos_map = {}
    excel_content = None
    excel_filename = ""

    if filename.endswith(".zip"):
        try:
            zf = zipfile.ZipFile(io.BytesIO(content))
        except zipfile.BadZipFile:
            raise HTTPException(status_code=400, detail="Fichier ZIP invalide")

        for name in zf.namelist():
            if name.startswith("__MACOSX") or name.startswith("."):
                continue
            lower = name.lower()
            basename = name.split("/")[-1]
            if not basename:
                continue
            if lower.endswith((".xlsx", ".xls", ".csv")):
                excel_content = zf.read(name)
                excel_filename = basename
            elif lower.endswith((".jpg", ".jpeg", ".png", ".webp")):
                photo_data = zf.read(name)
                ext = lower.rsplit(".", 1)[-1]
                mime = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", "webp": "image/webp"}[ext]
                b64 = base64.b64encode(photo_data).decode("utf-8")
                data_uri = f"data:{mime};base64,{b64}"
                photos_map[basename.lower()] = data_uri
                photos_map[basename.rsplit(".", 1)[0].lower()] = data_uri

        if not excel_content:
            raise HTTPException(status_code=400, detail="Aucun fichier Excel/CSV trouvé dans le ZIP")
        content = excel_content
        filename = excel_filename

    rows = []
    if filename.endswith(".csv"):
        import csv
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text), delimiter=";")
        first_row = next(csv.DictReader(io.StringIO(text), delimiter=";"), None)
        if first_row and len(first_row) <= 1:
            reader = csv.DictReader(io.StringIO(text), delimiter=",")
        else:
            reader = csv.DictReader(io.StringIO(text), delimiter=";")
        for row in reader:
            rows.append(row)
    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = str(val).strip() if val is not None else ""
            rows.append(row_dict)
    else:
        raise HTTPException(status_code=400, detail="Format non supporté. Utilisez .xlsx, .csv ou .zip")

    if not rows:
        raise HTTPException(status_code=400, detail="Le fichier est vide")

    def find_col(row, candidates):
        for c in candidates:
            for key in row.keys():
                if c in key.lower():
                    return row[key]
        return ""

    agency_id = user.get("agency_id")
    default_password = hash_password("LogiRent2024")
    created = 0
    skipped = 0
    photos_matched = 0
    errors = []

    for i, row in enumerate(rows):
        name = find_col(row, ["nom", "name", "prenom", "prénom", "client"])
        email = find_col(row, ["email", "mail", "e-mail", "courriel"])
        phone = find_col(row, ["tel", "téléphone", "telephone", "phone", "mobile", "portable"])
        address = find_col(row, ["adresse", "address", "ville", "city"])
        photo_ref = find_col(row, ["photo", "image", "avatar", "picture", "profil"])

        prenom = find_col(row, ["prenom", "prénom", "firstname", "first_name"])
        nom = find_col(row, ["nom", "lastname", "last_name", "family"])
        if prenom and nom and not name:
            name = f"{prenom} {nom}"
        elif nom and not name:
            name = nom

        if not email:
            errors.append(f"Ligne {i+2}: email manquant")
            continue

        email = email.lower().strip()
        existing = await db.users.find_one({"email": email})
        if existing:
            skipped += 1
            continue

        profile_photo = None
        if photo_ref and photos_map:
            ref_lower = photo_ref.lower().strip()
            profile_photo = photos_map.get(ref_lower) or photos_map.get(ref_lower.rsplit(".", 1)[0])
        if not profile_photo and photos_map:
            name_key = (name or "").lower().replace(" ", "_").replace(" ", "")
            email_key = email.split("@")[0].lower()
            for key in [name_key, email_key]:
                if key and key in photos_map:
                    profile_photo = photos_map[key]
                    break
        if profile_photo:
            photos_matched += 1

        new_user = {
            "id": str(uuid.uuid4()),
            "email": email,
            "password_hash": default_password,
            "name": name or email.split("@")[0],
            "phone": phone or None,
            "address": address or None,
            "id_photo": None,
            "license_photo": None,
            "profile_photo": profile_photo,
            "role": "client",
            "agency_id": agency_id,
            "created_at": datetime.utcnow(),
        }
        await db.users.insert_one(new_user)
        created += 1

    return {
        "message": f"Import terminé: {created} créés, {skipped} existants, {photos_matched} photos, {len(errors)} erreurs",
        "created": created, "skipped": skipped,
        "photos_matched": photos_matched, "errors": errors[:20],
    }


@router.get("/admin/reservations")
async def get_admin_reservations(skip: int = 0, limit: int = 20, status: Optional[str] = None, user: dict = Depends(get_admin_user)):
    agency_id = user.get('agency_id')
    is_super = user.get('role') == 'super_admin'

    query = {} if is_super else {"agency_id": agency_id}
    if status:
        query["status"] = status

    reservations = await db.reservations.find(query).sort("start_date", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.reservations.count_documents(query)

    # Batch fetch users and vehicles
    user_ids = list(set(r.get('user_id') for r in reservations if r.get('user_id')))
    vehicle_ids = list(set(r.get('vehicle_id') for r in reservations if r.get('vehicle_id')))
    users_list = await db.users.find({"id": {"$in": user_ids}}, {"_id": 0, "id": 1, "name": 1, "email": 1, "id_photo": 1, "license_photo": 1}).to_list(len(user_ids))
    vehicles_list = await db.vehicles.find({"id": {"$in": vehicle_ids}}, {"_id": 0, "id": 1, "brand": 1, "model": 1}).to_list(len(vehicle_ids))
    users_map = {u['id']: u for u in users_list}
    vehicles_map = {v['id']: v for v in vehicles_list}

    for res in reservations:
        res['_id'] = str(res['_id'])
        u = users_map.get(res.get('user_id'))
        v = vehicles_map.get(res.get('vehicle_id'))
        res['user_name'] = u['name'] if u else 'Unknown'
        res['user_email'] = u['email'] if u else 'Unknown'
        res['vehicle_name'] = f"{v['brand']} {v['model']}" if v else 'Unknown'
        res['docs_missing'] = not (u.get('id_photo') and u.get('license_photo')) if u else True

    return {"reservations": reservations, "total": total}


@router.put("/admin/reservations/{reservation_id}/status")
async def update_reservation_status(reservation_id: str, status: str, user: dict = Depends(get_admin_user)):
    valid_statuses = ["pending", "pending_cash", "confirmed", "active", "completed", "cancelled"]
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")

    result = await db.reservations.update_one(
        {"id": reservation_id},
        {"$set": {"status": status, "updated_at": datetime.utcnow()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Reservation not found")

    reservation = await db.reservations.find_one({"id": reservation_id})
    if reservation:
        client = await db.users.find_one({"id": reservation.get('user_id')})
        vehicle = await db.vehicles.find_one({"id": reservation.get('vehicle_id')})
        vname = f"{vehicle['brand']} {vehicle['model']}" if vehicle else "Véhicule"
        cname = client.get('name', 'Client') if client else 'Client'

        status_msgs = {
            'confirmed': f"Votre réservation pour {vname} a été confirmée.",
            'active': f"Votre réservation pour {vname} est maintenant active. Bon trajet !",
            'completed': f"Votre location de {vname} est terminée. Merci !",
            'cancelled': f"Votre réservation pour {vname} a été annulée.",
        }
        notif_types = {
            'confirmed': 'reservation_confirmed', 'active': 'reservation_active',
            'completed': 'reservation_completed', 'cancelled': 'reservation_cancelled',
        }
        if status in status_msgs and client:
            await create_notification(client['id'], notif_types[status], status_msgs[status], reservation_id)
            try:
                email_html = generate_status_change_email(cname, vname, status, reservation)
                email_subjects = {
                    'confirmed': f"Réservation confirmée - {vname}",
                    'active': f"Location en cours - {vname}",
                    'completed': f"Location terminée - {vname}",
                    'cancelled': f"Réservation annulée - {vname}",
                }
                await send_email(client['email'], email_subjects.get(status, f"Mise à jour réservation - {vname}"), email_html, agency_id=user.get('agency_id'))
            except Exception as e:
                logger.error(f"Failed to send status change email: {e}")

    return {"message": f"Reservation status updated to {status}"}


from models import VehicleReturnData

@router.post("/admin/reservations/{reservation_id}/return")
async def process_vehicle_return(reservation_id: str, data: VehicleReturnData, user: dict = Depends(get_admin_user)):
    reservation = await db.reservations.find_one({"id": reservation_id}, {"_id": 0})
    if not reservation:
        raise HTTPException(status_code=404, detail="Reservation not found")

    if reservation.get('status') not in ('active', 'confirmed'):
        raise HTTPException(status_code=400, detail="Only active/confirmed reservations can be returned")

    vehicle = await db.vehicles.find_one({"id": reservation['vehicle_id']}, {"_id": 0})
    client = await db.users.find_one({"id": reservation.get('user_id')}, {"_id": 0})
    vname = f"{vehicle['brand']} {vehicle['model']}" if vehicle else "Vehicule"
    cname = client.get('name', 'Client') if client else 'Client'

    total_days = reservation.get('total_days', 1)
    surcharges = []
    total_surcharge = 0.0

    # 0. Auto-calculate extra days: compare actual return (now) vs planned end_date
    planned_end = reservation.get('end_date')
    actual_return = datetime.utcnow()
    extra_days = 0
    extra_hours = 0
    extra_days_cost = 0.0
    if planned_end:
        if isinstance(planned_end, str):
            planned_end = datetime.fromisoformat(planned_end.replace('Z', '+00:00').replace('+00:00', ''))
        diff_hours = (actual_return - planned_end).total_seconds() / 3600
        if diff_hours > 0:
            # Full 24h blocks = extra days, remaining = extra hours
            extra_days = int(diff_hours // 24)
            extra_hours = round(diff_hours % 24, 1)
            price_per_day = vehicle.get('price_per_day', 0) if vehicle else 0
            if extra_days > 0:
                extra_days_cost = round(extra_days * price_per_day, 2)
                surcharges.append({
                    "type": "extra_days",
                    "label": f"Jours supplementaires ({extra_days}j x CHF {price_per_day})",
                    "amount": extra_days_cost,
                    "editable": True
                })
                total_surcharge += extra_days_cost
            if extra_hours > 0:
                extra_hours_cost = round(extra_hours * data.late_penalty_per_hour, 2)
                surcharges.append({
                    "type": "extra_hours",
                    "label": f"Retard ({extra_hours}h x CHF {data.late_penalty_per_hour}/h)",
                    "amount": extra_hours_cost,
                    "editable": True
                })
                total_surcharge += extra_hours_cost

    # 1. Kilometrage
    km_driven = max(0, data.km_return - data.km_departure)
    km_allowed = data.km_limit_per_day * total_days
    km_excess = max(0, km_driven - km_allowed)
    if km_excess > 0:
        km_cost = round(km_excess * data.price_per_extra_km, 2)
        surcharges.append({"type": "km_excess", "label": f"Depassement km ({km_excess} km x CHF {data.price_per_extra_km})", "amount": km_cost})
        total_surcharge += km_cost

    # 2. Carburant
    fuel_levels = {"full": 1.0, "3/4": 0.75, "1/2": 0.5, "1/4": 0.25, "empty": 0.0}
    fuel_dep = fuel_levels.get(data.fuel_level_departure, 1.0)
    fuel_ret = fuel_levels.get(data.fuel_level_return, 1.0)
    if fuel_ret < fuel_dep:
        fuel_diff = fuel_dep - fuel_ret
        fuel_cost = round(data.fuel_penalty * fuel_diff, 2)
        surcharges.append({"type": "fuel", "label": f"Carburant manquant ({data.fuel_level_departure} -> {data.fuel_level_return})", "amount": fuel_cost})
        total_surcharge += fuel_cost

    # 3. Retard
    if data.late_hours > 0:
        late_cost = round(data.late_hours * data.late_penalty_per_hour, 2)
        surcharges.append({"type": "late", "label": f"Retard ({data.late_hours}h x CHF {data.late_penalty_per_hour})", "amount": late_cost})
        total_surcharge += late_cost

    # Save return data
    return_record = {
        "id": str(uuid.uuid4()),
        "reservation_id": reservation_id,
        "vehicle_id": reservation['vehicle_id'],
        "km_departure": data.km_departure,
        "km_return": data.km_return,
        "km_driven": km_driven,
        "km_allowed": km_allowed,
        "km_excess": km_excess,
        "fuel_departure": data.fuel_level_departure,
        "fuel_return": data.fuel_level_return,
        "late_hours": data.late_hours,
        "extra_days": extra_days,
        "extra_days_cost": extra_days_cost,
        "surcharges": surcharges,
        "total_surcharge": total_surcharge,
        "notes": data.notes,
        "processed_by": user.get('id'),
        "processed_at": datetime.utcnow().isoformat(),
    }
    await db.vehicle_returns.insert_one(return_record)

    # Update reservation: completed + add surcharge
    update_data = {
        "status": "completed",
        "updated_at": datetime.utcnow(),
        "return_data": {
            "km_driven": km_driven,
            "km_excess": km_excess,
            "fuel_return": data.fuel_level_return,
            "late_hours": data.late_hours,
            "extra_days": extra_days,
            "extra_days_cost": extra_days_cost,
            "surcharges": surcharges,
            "total_surcharge": total_surcharge,
            "processed_at": datetime.utcnow().isoformat(),
        },
        "actual_return_date": datetime.utcnow(),
    }
    if total_surcharge > 0:
        update_data["total_price"] = reservation.get('total_price', 0) + total_surcharge

    await db.reservations.update_one({"id": reservation_id}, {"$set": update_data})

    # Update vehicle status back to available
    await db.vehicles.update_one({"id": reservation['vehicle_id']}, {"$set": {"status": "available"}})

    # Notify client
    msg = f"Votre location de {vname} est terminee. "
    if total_surcharge > 0:
        msg += f"Supplements: CHF {total_surcharge:.2f}."
    else:
        msg += "Aucun supplement."
    if client:
        await create_notification(client['id'], 'reservation_completed', msg, reservation_id)

    return {
        "message": "Retour enregistre",
        "km_driven": km_driven,
        "km_excess": km_excess,
        "extra_days": extra_days,
        "extra_days_cost": extra_days_cost,
        "surcharges": surcharges,
        "total_surcharge": total_surcharge,
        "new_total": reservation.get('total_price', 0) + total_surcharge,
    }


@router.put("/admin/reservations/{reservation_id}/adjust-price")
async def adjust_reservation_price(reservation_id: str, request: Request, user: dict = Depends(get_admin_user)):
    """Admin manually adjusts the total price of a reservation"""
    body = await request.json()
    new_total = body.get('total_price')
    reason = body.get('reason', '')
    if new_total is None or new_total < 0:
        raise HTTPException(status_code=400, detail="Prix invalide")
    res = await db.reservations.find_one({"id": reservation_id}, {"_id": 0})
    if not res:
        raise HTTPException(status_code=404, detail="Reservation non trouvee")
    old_total = res.get('total_price', 0)
    await db.reservations.update_one({"id": reservation_id}, {"$set": {
        "total_price": float(new_total),
        "updated_at": datetime.utcnow(),
    }, "$push": {
        "price_adjustments": {
            "old_price": old_total,
            "new_price": float(new_total),
            "reason": reason,
            "adjusted_by": user.get('id'),
            "adjusted_at": datetime.utcnow().isoformat(),
        }
    }})
    return {"message": f"Prix ajuste de CHF {old_total:.2f} a CHF {float(new_total):.2f}", "old_total": old_total, "new_total": float(new_total)}


@router.post("/admin/reservations/{reservation_id}/send-offer")
async def send_price_offer(reservation_id: str, request: Request, user: dict = Depends(get_admin_user)):
    """Admin sends a (re)priced offer to the client for a pending reservation.
    Optionally updates the total price and an accompanying message, then emails the client."""
    from utils.email import send_price_offer_email
    from utils.notifications import create_notification

    body = await request.json()
    new_total = body.get('total_price')
    message = (body.get('message') or '').strip()

    res = await db.reservations.find_one({"id": reservation_id}, {"_id": 0})
    if not res:
        raise HTTPException(status_code=404, detail="Reservation non trouvee")
    if res.get('status') not in ['pending', 'pending_cash']:
        raise HTTPException(status_code=400, detail="L'offre ne peut etre envoyee que pour une demande en attente")

    old_total = float(res.get('total_price', 0))
    update: dict = {"updated_at": datetime.utcnow(), "last_offer_sent_at": datetime.utcnow()}
    push: dict = {}
    price_changed = False

    if new_total is not None:
        try:
            new_total = float(new_total)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Prix invalide")
        if new_total < 0:
            raise HTTPException(status_code=400, detail="Prix invalide")
        if abs(new_total - old_total) > 0.01:
            update["total_price"] = new_total
            push["price_adjustments"] = {
                "old_price": old_total,
                "new_price": new_total,
                "reason": message or "Offre ajustee par l'agence",
                "adjusted_by": user.get('id'),
                "adjusted_at": datetime.utcnow().isoformat(),
            }
            price_changed = True
    else:
        new_total = old_total

    update_ops: dict = {"$set": update}
    if push:
        update_ops["$push"] = push
    await db.reservations.update_one({"id": reservation_id}, update_ops)

    # Fetch related entities for email
    client = await db.users.find_one({"id": res.get('user_id')}, {"_id": 0})
    vehicle = await db.vehicles.find_one({"id": res.get('vehicle_id')}, {"_id": 0})
    reservation_updated = await db.reservations.find_one({"id": reservation_id}, {"_id": 0})

    if client and vehicle:
        try:
            await send_price_offer_email(
                client, vehicle, reservation_updated,
                old_price=old_total, new_price=new_total,
                message=message, agency_id=res.get('agency_id')
            )
        except Exception as e:
            logger.error(f"Failed to send offer email: {e}")

        try:
            vname = f"{vehicle.get('brand', '')} {vehicle.get('model', '')}".strip()
            if price_changed:
                msg = f"L'agence a revu le prix de votre demande pour {vname}: CHF {old_total:.2f} -> CHF {new_total:.2f}. Consultez votre e-mail pour accepter l'offre."
            else:
                msg = f"L'agence vous a envoye une offre pour {vname} (CHF {new_total:.2f}). Consultez votre e-mail pour accepter l'offre."
            await create_notification(res.get('user_id'), 'status_changed', msg, reservation_id)
        except Exception as e:
            logger.error(f"Failed to notify client of offer: {e}")

    return {
        "message": "Offre envoyee au client",
        "old_total": old_total,
        "new_total": new_total,
        "price_changed": price_changed,
    }



@router.put("/admin/reservations/{reservation_id}/payment-status")
async def update_payment_status(reservation_id: str, payment_status: str, user: dict = Depends(get_admin_user)):
    valid_statuses = ["unpaid", "pending", "paid", "refunded"]
    if payment_status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid payment status. Must be one of: {valid_statuses}")

    reservation = await db.reservations.find_one({"id": reservation_id})
    if not reservation:
        raise HTTPException(status_code=404, detail="Reservation not found")

    update_data = {"payment_status": payment_status, "updated_at": datetime.utcnow()}
    if payment_status == "paid" and reservation.get('status') in ['pending_cash', 'pending']:
        update_data["status"] = "confirmed"

    await db.reservations.update_one({"id": reservation_id}, {"$set": update_data})

    if payment_status == "paid":
        existing_tx = await db.payment_transactions.find_one({"reservation_id": reservation_id})
        if existing_tx:
            await db.payment_transactions.update_one(
                {"reservation_id": reservation_id},
                {"$set": {"status": "paid", "payment_status": "paid", "updated_at": datetime.utcnow()}}
            )
        else:
            new_transaction = PaymentTransaction(
                user_id=reservation['user_id'], reservation_id=reservation_id,
                session_id=f"cash_{reservation_id}", amount=float(reservation['total_price']),
                currency="chf", status="paid", payment_status="paid",
                metadata={"payment_method": reservation.get('payment_method', 'cash'), "admin_confirmed": True}
            )
            await db.payment_transactions.insert_one(new_transaction.dict())

        try:
            res_user = await db.users.find_one({"id": reservation['user_id']})
            vehicle = await db.vehicles.find_one({"id": reservation['vehicle_id']})
            if res_user and vehicle:
                await send_payment_confirmation(res_user, vehicle, reservation)
                # Notify client: payment confirmed by admin
                vname = f"{vehicle['brand']} {vehicle['model']}"
                await create_notification(
                    res_user['id'], 'payment_success',
                    f"Votre paiement de CHF {reservation['total_price']:.2f} pour {vname} a ete confirme. Votre reservation est entierement validee.",
                    reservation_id
                )
        except Exception as email_error:
            logger.error(f"Failed to send confirmation email: {email_error}")

    return {"message": f"Payment status updated to {payment_status}"}




@router.put("/admin/reservations/{reservation_id}/reschedule")
async def reschedule_reservation(reservation_id: str, new_start: str, new_end: str, user: dict = Depends(get_admin_user)):
    reservation = await db.reservations.find_one({"id": reservation_id})
    if not reservation:
        raise HTTPException(status_code=404, detail="Reservation introuvable")
    if reservation['status'] in ['completed', 'cancelled']:
        raise HTTPException(status_code=400, detail="Impossible de deplacer une reservation terminee ou annulee")

    start_dt = datetime.fromisoformat(new_start)
    end_dt = datetime.fromisoformat(new_end)
    total_days = (end_dt - start_dt).days
    if total_days <= 0:
        raise HTTPException(status_code=400, detail="La date de fin doit etre apres la date de debut")

    overlap = await db.reservations.find_one({
        "vehicle_id": reservation['vehicle_id'],
        "id": {"$ne": reservation_id},
        "status": {"$nin": ["cancelled", "completed"]},
        "start_date": {"$lt": end_dt},
        "end_date": {"$gt": start_dt},
    })
    if overlap:
        raise HTTPException(status_code=409, detail="Conflit: un autre vehicule est reserve sur cette periode")

    vehicle = await db.vehicles.find_one({"id": reservation['vehicle_id']})
    base_price = vehicle['price_per_day'] * total_days
    options_price = sum(opt.get('price_per_day', 0) * total_days for opt in reservation.get('options', []))
    total_price = base_price + options_price

    await db.reservations.update_one({"id": reservation_id}, {"$set": {
        "start_date": start_dt, "end_date": end_dt,
        "total_days": total_days, "base_price": base_price,
        "options_price": options_price, "total_price": total_price,
        "updated_at": datetime.utcnow(),
    }})
    return {"success": True, "message": f"Reservation deplacee du {new_start} au {new_end}", "total_days": total_days, "total_price": total_price}


@router.get("/admin/payments")
async def get_admin_payments(skip: int = 0, limit: int = 20, user: dict = Depends(get_admin_user)):
    transactions = await db.payment_transactions.find({}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.payment_transactions.count_documents({})

    # Batch fetch users
    tx_user_ids = list(set(tx.get('user_id') for tx in transactions if tx.get('user_id')))
    tx_users = await db.users.find({"id": {"$in": tx_user_ids}}, {"_id": 0, "id": 1, "email": 1}).to_list(len(tx_user_ids))
    tx_users_map = {u['id']: u for u in tx_users}

    for tx in transactions:
        tx['_id'] = str(tx['_id'])
        u = tx_users_map.get(tx.get('user_id'))
        tx['user_email'] = u['email'] if u else 'Unknown'

    return {"transactions": transactions, "total": total}


@router.get("/admin/calendar")
async def get_admin_calendar(month: int = None, year: int = None, user: dict = Depends(get_admin_user)):
    agency_id = user.get('agency_id')
    is_super = user.get('role') == 'super_admin'

    now = datetime.utcnow()
    if month is None:
        month = now.month
    if year is None:
        year = now.year

    start_of_month = datetime(year, month, 1)
    end_of_month = datetime(year + 1, 1, 1) if month == 12 else datetime(year, month + 1, 1)

    cal_query = {
        "status": {"$in": ["pending", "pending_cash", "confirmed", "active", "completed"]},
        "start_date": {"$lt": end_of_month}, "end_date": {"$gt": start_of_month}
    }
    if not is_super:
        cal_query["agency_id"] = agency_id

    reservations = await db.reservations.find(cal_query).to_list(500)

    # Batch fetch users and vehicles
    cal_user_ids = list(set(r.get('user_id') for r in reservations if r.get('user_id')))
    cal_vehicle_ids = list(set(r.get('vehicle_id') for r in reservations if r.get('vehicle_id')))
    cal_users = await db.users.find({"id": {"$in": cal_user_ids}}, {"_id": 0, "id": 1, "name": 1, "email": 1, "phone": 1}).to_list(len(cal_user_ids))
    cal_vehicles = await db.vehicles.find({"id": {"$in": cal_vehicle_ids}}, {"_id": 0, "id": 1, "brand": 1, "model": 1}).to_list(len(cal_vehicle_ids))
    cal_users_map = {u['id']: u for u in cal_users}
    cal_vehicles_map = {v['id']: v for v in cal_vehicles}

    events = []
    for res in reservations:
        u = cal_users_map.get(res.get('user_id'))
        v = cal_vehicles_map.get(res.get('vehicle_id'))
        is_overdue = res['status'] == 'active' and res['end_date'] < now

        events.append({
            "id": res['id'],
            "user_name": u['name'] if u else 'Inconnu',
            "user_email": u['email'] if u else '',
            "user_phone": u.get('phone', '') if u else '',
            "vehicle_name": f"{v['brand']} {v['model']}" if v else 'Inconnu',
            "vehicle_id": res.get('vehicle_id', ''),
            "start_date": res['start_date'].isoformat(),
            "end_date": res['end_date'].isoformat(),
            "total_days": res.get('total_days', 0),
            "total_price": res.get('total_price', 0),
            "status": res['status'],
            "payment_status": res.get('payment_status', 'unpaid'),
            "payment_method": res.get('payment_method', 'card'),
            "is_overdue": is_overdue,
            "days_overdue": (now - res['end_date']).days if is_overdue else 0
        })

    return {"events": events, "month": month, "year": year}


@router.get("/admin/overdue")
async def get_overdue_reservations(user: dict = Depends(get_admin_user)):
    now = datetime.utcnow()
    overdue_reservations = await db.reservations.find({"status": "active", "end_date": {"$lt": now}}).sort("end_date", 1).to_list(100)

    # Batch fetch users and vehicles
    od_user_ids = list(set(r.get('user_id') for r in overdue_reservations if r.get('user_id')))
    od_vehicle_ids = list(set(r.get('vehicle_id') for r in overdue_reservations if r.get('vehicle_id')))
    od_users = await db.users.find({"id": {"$in": od_user_ids}}, {"_id": 0, "id": 1, "name": 1, "email": 1, "phone": 1}).to_list(len(od_user_ids))
    od_vehicles = await db.vehicles.find({"id": {"$in": od_vehicle_ids}}, {"_id": 0, "id": 1, "brand": 1, "model": 1}).to_list(len(od_vehicle_ids))
    od_users_map = {u['id']: u for u in od_users}
    od_vehicles_map = {v['id']: v for v in od_vehicles}

    results = []
    for res in overdue_reservations:
        u = od_users_map.get(res.get('user_id'))
        v = od_vehicles_map.get(res.get('vehicle_id'))
        results.append({
            "id": res['id'],
            "user_name": u['name'] if u else 'Inconnu',
            "user_email": u['email'] if u else '',
            "user_phone": u.get('phone', '') if u else '',
            "vehicle_name": f"{v['brand']} {v['model']}" if v else 'Inconnu',
            "start_date": res['start_date'].isoformat(),
            "end_date": res['end_date'].isoformat(),
            "total_days": res.get('total_days', 0),
            "total_price": res.get('total_price', 0),
            "days_overdue": (now - res['end_date']).days,
            "status": res['status'],
            "payment_status": res.get('payment_status', 'unpaid'),
        })

    return {"overdue": results, "total": len(results)}


@router.put("/admin/vehicles/{vehicle_id}/status")
async def update_vehicle_status(vehicle_id: str, status: str, user: dict = Depends(get_admin_user)):
    valid_statuses = ["available", "rented", "maintenance"]
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")

    result = await db.vehicles.update_one({"id": vehicle_id}, {"$set": {"status": status}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    return {"message": f"Vehicle status updated to {status}"}


# Background cron task
async def _check_overdue_task():
    now = datetime.utcnow()
    overdue_reservations = await db.reservations.find({"status": "active", "end_date": {"$lt": now}}).to_list(100)

    created = 0
    for res in overdue_reservations:
        existing = await db.notifications.find_one({"reservation_id": res['id'], "type": "late_return"})
        if existing:
            continue

        vehicle = await db.vehicles.find_one({"id": res['vehicle_id']})
        vehicle_name = f"{vehicle['brand']} {vehicle['model']}" if vehicle else 'Véhicule'
        days_overdue = (now - res['end_date']).days

        notification = {
            "id": str(uuid.uuid4()),
            "user_id": res['user_id'],
            "reservation_id": res['id'],
            "type": "late_return",
            "title": "Retour en retard",
            "message": f"Votre location de {vehicle_name} est en retard de {days_overdue} jour(s). Veuillez retourner le véhicule dès que possible.",
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(notification)
        created += 1

        try:
            res_user = await db.users.find_one({"id": res['user_id']})
            if res_user:
                html = f"""
                <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:20px;">
                    <div style="background:#EF4444;padding:20px;border-radius:12px 12px 0 0;text-align:center;">
                        <h2 style="color:#fff;margin:0;">Retour en retard</h2>
                    </div>
                    <div style="background:#fff;padding:20px;border-radius:0 0 12px 12px;border:1px solid #E5E7EB;">
                        <p>Bonjour {res_user['name']},</p>
                        <p>Votre location de <strong>{vehicle_name}</strong> devait être retournée le <strong>{res['end_date'].strftime('%d/%m/%Y')}</strong>.</p>
                        <p style="color:#EF4444;font-weight:bold;">Vous avez {days_overdue} jour(s) de retard.</p>
                        <p>Veuillez retourner le véhicule dès que possible.</p>
                        <p>L'équipe LogiRent</p>
                    </div>
                </div>
                """
                await send_email(res_user['email'], f"Retour en retard - {vehicle_name}", html)
        except Exception as e:
            logger.error(f"Failed to send overdue email: {e}")

    return created


async def overdue_cron_loop():
    while True:
        try:
            created = await _check_overdue_task()
            if created > 0:
                logger.info(f"Overdue cron: created {created} late return notifications")
        except Exception as e:
            logger.error(f"Overdue cron error: {e}")
        await asyncio.sleep(3600)


@router.post("/admin/check-overdue")
async def check_overdue_and_notify(user: dict = Depends(get_admin_user)):
    created = await _check_overdue_task()
    return {"message": f"Checked overdue reservations. Created {created} new notifications."}



@router.put("/admin/agencies/{agency_id}/toggle-active")
async def toggle_agency_active(agency_id: str, user: dict = Depends(get_admin_user)):
    if user.get('role') != 'super_admin':
        raise HTTPException(status_code=403, detail="Super admin only")
    
    agency = await db.agencies.find_one({"id": agency_id})
    if not agency:
        raise HTTPException(status_code=404, detail="Agency not found")
    
    current_status = agency.get('is_active', True)
    new_status = not current_status
    
    # Update agency status
    await db.agencies.update_one({"id": agency_id}, {"$set": {"is_active": new_status}})
    
    # Also update all admin users of this agency
    await db.users.update_many(
        {"agency_id": agency_id, "role": "admin"},
        {"$set": {"is_active": new_status}}
    )
    
    status_text = "activé" if new_status else "désactivé"
    return {"message": f"Agence {status_text}", "is_active": new_status}



# ======================== AGENCY MODULES ========================

DEFAULT_MODULES = {
    "online_booking": True,
    "stripe_payment": True,
    "cash_payment": True,
    "inspections": True,
    "ai_damage": True,
    "email_notifications": True,
    "gps_tracking": True,
    "e_signature": True,
    "analytics": True,
    "vehicle_return": True,
    "availability_calendar": True,
    "smtp_custom": True,
    "qr_invoicing": True,
    "gantt_planning": True,
    "document_scan": True,
    "predictive_maintenance": False,
    "dynamic_pricing": False,
}

MODULE_LABELS = {
    "online_booking": "Reservations en ligne",
    "stripe_payment": "Paiement Stripe",
    "cash_payment": "Paiement en especes",
    "inspections": "Inspections vehicules",
    "ai_damage": "Detection de dommages IA",
    "email_notifications": "Notifications email",
    "gps_tracking": "GPS / Tracking",
    "e_signature": "E-Signature contrats",
    "analytics": "Analytics Dashboard",
    "vehicle_return": "Gestion de retour vehicule",
    "availability_calendar": "Calendrier de disponibilite",
    "smtp_custom": "Email SMTP personnalise",
    "qr_invoicing": "Facturation QR Swiss",
    "gantt_planning": "Planning Gantt interactif",
    "document_scan": "Scan documents IA (OCR)",
    "predictive_maintenance": "Maintenance predictive",
    "dynamic_pricing": "Tarification dynamique IA",
}


@router.get("/admin/agencies/{agency_id}/modules")
async def get_agency_modules(agency_id: str, user: dict = Depends(get_admin_user)):
    if user.get('role') != 'super_admin':
        raise HTTPException(status_code=403, detail="Super admin only")
    agency = await db.agencies.find_one({"id": agency_id}, {"_id": 0})
    if not agency:
        raise HTTPException(status_code=404, detail="Agency not found")
    modules = agency.get('modules', DEFAULT_MODULES.copy())
    # Ensure all default modules exist
    for key, val in DEFAULT_MODULES.items():
        if key not in modules:
            modules[key] = val
    return {"agency_id": agency_id, "modules": modules, "labels": MODULE_LABELS}


@router.put("/admin/agencies/{agency_id}/modules")
async def update_agency_modules(agency_id: str, data: dict, user: dict = Depends(get_admin_user)):
    if user.get('role') != 'super_admin':
        raise HTTPException(status_code=403, detail="Super admin only")
    agency = await db.agencies.find_one({"id": agency_id})
    if not agency:
        raise HTTPException(status_code=404, detail="Agency not found")
    modules = data.get('modules', {})
    # Only allow valid module keys
    clean_modules = {}
    for key in DEFAULT_MODULES:
        clean_modules[key] = bool(modules.get(key, DEFAULT_MODULES[key]))
    await db.agencies.update_one({"id": agency_id}, {"$set": {"modules": clean_modules}})
    return {"message": "Modules mis a jour", "modules": clean_modules}


@router.get("/agency-modules")
async def get_my_agency_modules(user: dict = Depends(get_admin_user)):
    """Get modules for the current user's agency"""
    agency_id = user.get('agency_id')
    if not agency_id:
        return {"modules": DEFAULT_MODULES.copy()}
    agency = await db.agencies.find_one({"id": agency_id}, {"_id": 0})
    if not agency:
        return {"modules": DEFAULT_MODULES.copy()}
    modules = agency.get('modules', DEFAULT_MODULES.copy())
    for key, val in DEFAULT_MODULES.items():
        if key not in modules:
            modules[key] = val
    return {"modules": modules}


# ======================== AGENCY BILLING SETTINGS ========================

class BillingSettings(PydanticBaseModel):
    company_name: str = ""
    street: str = ""
    house_number: str = ""
    pcode: str = ""
    city: str = ""
    country: str = "CH"
    phone: str = ""
    email: str = ""
    website: str = ""
    iban: str = ""
    vat_number: str = ""


@router.get("/admin/billing-settings")
async def get_billing_settings(user: dict = Depends(get_agency_admin)):
    """Get agency billing settings (IBAN, company info for invoices)."""
    agency_id = user.get('agency_id')
    settings = await db.billing_settings.find_one({"agency_id": agency_id}, {"_id": 0})
    if not settings:
        return {
            "agency_id": agency_id,
            "company_name": "",
            "street": "",
            "house_number": "",
            "pcode": "",
            "city": "",
            "country": "CH",
            "phone": "",
            "email": "",
            "website": "",
            "iban": "",
            "vat_number": "",
        }
    return settings


@router.put("/admin/billing-settings")
async def update_billing_settings(data: BillingSettings, user: dict = Depends(get_agency_admin)):
    """Update agency billing settings (IBAN, company info for invoices)."""
    agency_id = user.get('agency_id')
    doc = {
        "agency_id": agency_id,
        "company_name": data.company_name,
        "street": data.street,
        "house_number": data.house_number,
        "pcode": data.pcode,
        "city": data.city,
        "country": data.country,
        "phone": data.phone,
        "email": data.email,
        "website": data.website,
        "iban": data.iban,
        "vat_number": data.vat_number,
        "updated_at": datetime.utcnow(),
    }
    await db.billing_settings.update_one(
        {"agency_id": agency_id},
        {"$set": doc},
        upsert=True,
    )
    return {"message": "Parametres de facturation mis a jour", **doc}


# ======================== DOCUMENT SCANNING ========================

@router.post("/admin/clients/{client_id}/documents")
async def upload_client_document(client_id: str, user: dict = Depends(get_agency_admin)):
    """Upload a client document (ID card, driver license) via base64."""
    from fastapi import Request
    # This endpoint handles base64 uploads from the camera/gallery
    pass


@router.post("/documents/upload")
async def upload_document(
    doc_type: str,
    file: UploadFile = File(...),
    user: dict = Depends(get_admin_user),
):
    """Upload a scanned document (ID card, license) to secure storage."""
    from utils.storage import put_object, get_public_url

    allowed_types = ['id_card_front', 'id_card_back', 'license_front', 'license_back', 'other']
    if doc_type not in allowed_types:
        raise HTTPException(status_code=400, detail=f"Type invalide. Types: {allowed_types}")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Fichier trop volumineux (max 10MB)")

    ext = file.filename.rsplit(".", 1)[-1].lower() if file.filename and "." in file.filename else "jpg"
    path = f"logirent/documents/{user.get('id', 'unknown')}/{doc_type}_{uuid.uuid4()}.{ext}"

    put_object(path, content, file.content_type or "image/jpeg")
    url = get_public_url(path)

    doc_record = {
        "id": str(uuid.uuid4()),
        "uploader_id": user['id'],
        "doc_type": doc_type,
        "filename": file.filename,
        "storage_path": path,
        "url": url,
        "status": "pending",  # pending, validated, rejected
        "extracted_data": {},
        "validated_by": None,
        "validated_at": None,
        "created_at": datetime.utcnow(),
    }
    await db.documents.insert_one(doc_record)
    doc_record.pop('_id', None)
    return doc_record


@router.post("/documents/upload-base64")
async def upload_document_base64(
    data: dict,
    user=Depends(get_current_user),
):
    """Upload a document from camera capture (base64 encoded)."""
    from utils.storage import put_object, get_public_url

    base64_data = data.get('image')
    doc_type = data.get('doc_type', 'other')
    client_id = data.get('client_id')
    filename = data.get('filename', 'capture.jpg')

    if not base64_data:
        raise HTTPException(status_code=400, detail="Image base64 manquante")

    # Strip data URI prefix if present
    if ',' in base64_data:
        base64_data = base64_data.split(',', 1)[1]

    import base64 as b64
    try:
        content = b64.b64decode(base64_data)
    except Exception:
        raise HTTPException(status_code=400, detail="Base64 invalide")

    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Fichier trop volumineux (max 10MB)")

    target_id = client_id or user['id']
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "jpg"
    path = f"logirent/documents/{target_id}/{doc_type}_{uuid.uuid4()}.{ext}"

    put_object(path, content, "image/jpeg")
    url = get_public_url(path)

    doc_record = {
        "id": str(uuid.uuid4()),
        "client_id": target_id,
        "uploader_id": user['id'],
        "doc_type": doc_type,
        "filename": filename,
        "storage_path": path,
        "url": url,
        "status": "pending",
        "extracted_data": {},
        "ocr_status": "processing",
        "ocr_confidence": 0,
        "validated_by": None,
        "validated_at": None,
        "created_at": datetime.utcnow(),
    }
    await db.documents.insert_one(doc_record)
    doc_id = doc_record["id"]
    doc_record.pop('_id', None)

    # Launch OCR in background
    asyncio.create_task(_run_ocr_background(doc_id, base64_data, doc_type))

    return doc_record


async def _run_ocr_background(doc_id: str, base64_data: str, doc_type: str):
    """Run OCR extraction in background after document upload."""
    try:
        ocr_result = await extract_document_ocr(base64_data, doc_type)
        update = {
            "ocr_status": "completed" if ocr_result.get("success") else "failed",
            "ocr_confidence": ocr_result.get("confidence", 0),
            "ocr_result": ocr_result,
        }
        if ocr_result.get("success") and ocr_result.get("extracted_data"):
            update["extracted_data"] = ocr_result["extracted_data"]
        await db.documents.update_one({"id": doc_id}, {"$set": update})
        logger.info(f"OCR completed for document {doc_id}: confidence={ocr_result.get('confidence', 0)}%")
    except Exception as e:
        logger.error(f"OCR background task failed for {doc_id}: {e}")
        await db.documents.update_one({"id": doc_id}, {"$set": {"ocr_status": "failed"}})


@router.get("/documents/client/{client_id}")
async def get_client_documents(client_id: str, user: dict = Depends(get_admin_user)):
    """List all documents for a client."""
    docs = await db.documents.find({"client_id": client_id}, {"_id": 0}).sort("created_at", -1).to_list(50)
    return docs


@router.get("/documents/my")
async def get_my_documents(user: dict = Depends(get_current_user)):
    """Get documents uploaded by or for the current user."""
    user_id = user['id']
    docs = await db.documents.find(
        {"$or": [{"client_id": user_id}, {"uploader_id": user_id}]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    return docs


@router.put("/documents/{doc_id}/validate")
async def validate_document(doc_id: str, data: dict, user: dict = Depends(get_agency_admin)):
    """Validate a document and save extracted data."""
    doc = await db.documents.find_one({"id": doc_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Document introuvable")

    extracted = data.get('extracted_data', {})
    status = data.get('status', 'validated')

    await db.documents.update_one(
        {"id": doc_id},
        {"$set": {
            "status": status,
            "extracted_data": extracted,
            "validated_by": user['id'],
            "validated_at": datetime.utcnow(),
        }}
    )

    # Update client profile with extracted data if available
    client_id = doc.get('client_id')
    if client_id and extracted and status == 'validated':
        update_fields = {}
        if extracted.get('name'):
            update_fields['name'] = extracted['name']
        if extracted.get('date_of_birth'):
            update_fields['date_of_birth'] = extracted['date_of_birth']
        if extracted.get('license_number'):
            update_fields['license_number'] = extracted['license_number']
        if extracted.get('license_expiry_date'):
            update_fields['license_expiry_date'] = extracted['license_expiry_date']
        if extracted.get('nationality'):
            update_fields['nationality'] = extracted['nationality']
        if update_fields:
            await db.users.update_one({"id": client_id}, {"$set": update_fields})

    return {"message": "Document valide", "status": status}


@router.post("/documents/{doc_id}/ocr")
async def trigger_document_ocr(doc_id: str, user: dict = Depends(get_admin_user)):
    """Manually trigger OCR extraction on a document."""
    doc = await db.documents.find_one({"id": doc_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Document introuvable")

    url = doc.get("url")
    if not url:
        raise HTTPException(status_code=400, detail="Document sans image")

    # Download image from URL and convert to base64
    import httpx
    import base64 as b64
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.get(url)
            resp.raise_for_status()
            image_b64 = b64.b64encode(resp.content).decode("utf-8")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Impossible de telecharger l'image: {str(e)[:100]}")

    await db.documents.update_one({"id": doc_id}, {"$set": {"ocr_status": "processing"}})

    ocr_result = await extract_document_ocr(image_b64, doc.get("doc_type", "other"))

    update = {
        "ocr_status": "completed" if ocr_result.get("success") else "failed",
        "ocr_confidence": ocr_result.get("confidence", 0),
        "ocr_result": ocr_result,
    }
    if ocr_result.get("success") and ocr_result.get("extracted_data"):
        update["extracted_data"] = ocr_result["extracted_data"]

    await db.documents.update_one({"id": doc_id}, {"$set": update})

    return {
        "message": "OCR termine",
        "ocr_status": update["ocr_status"],
        "confidence": update.get("ocr_confidence", 0),
        "extracted_data": ocr_result.get("extracted_data", {}),
    }


# ======================== CONTRACT TEMPLATE ========================

@router.get("/admin/contract-template")
async def get_contract_template(user: dict = Depends(get_agency_admin)):
    agency_id = user.get("agency_id")
    template = await db.contract_templates.find_one({"agency_id": agency_id}, {"_id": 0})
    if not template:
        # Return default template
        return {
            "agency_id": agency_id,
            "legal_text": (
                "Le/la soussigné(e) déclare avoir pris connaissance et accepter les conditions générales "
                "de location disponibles sur le site {website}, lesquelles font partie intégrante du présent document.\n\n"
                "Le locataire s'engage à utiliser le véhicule avec diligence et à respecter strictement les dispositions "
                "de la Loi fédérale sur la circulation routière (LCR) ainsi que toutes les prescriptions légales applicables.\n\n"
                "Les dommages couverts par l'assurance Casco collision du loueur sont soumis à une franchise de "
                "CHF {franchise}.– par sinistre, laquelle demeure entièrement à la charge du locataire ou du "
                "conducteur responsable.\n\n"
                "Le locataire reconnaît être responsable de tout dommage, amende ou frais résultant de l'utilisation "
                "du véhicule. Le présent document vaut reconnaissance de dette au sens de l'art. 82 LP."
            ),
            "default_prices": {},
            "deductible": "1000",
            "agency_website": "",
            "logo_path": None,
        }
    return template


@router.put("/admin/contract-template")
async def update_contract_template(data: dict, user: dict = Depends(get_agency_admin)):
    agency_id = user.get("agency_id")

    allowed = {"legal_text", "default_prices", "deductible", "agency_website"}
    update = {k: v for k, v in data.items() if k in allowed}
    update["updated_at"] = datetime.utcnow().isoformat()

    existing = await db.contract_templates.find_one({"agency_id": agency_id})
    if existing:
        await db.contract_templates.update_one(
            {"agency_id": agency_id}, {"$set": update}
        )
    else:
        update["agency_id"] = agency_id
        update["id"] = str(uuid.uuid4())
        update["logo_path"] = None
        update["created_at"] = datetime.utcnow().isoformat()
        await db.contract_templates.insert_one(update)

    template = await db.contract_templates.find_one({"agency_id": agency_id}, {"_id": 0})
    return template


@router.post("/admin/contract-template/logo")
async def upload_template_logo(file: UploadFile = File(...), user: dict = Depends(get_agency_admin)):
    from utils.storage import put_object
    agency_id = user.get("agency_id")

    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Logo trop volumineux (max 5MB)")

    ext = file.filename.split(".")[-1].lower() if file.filename and "." in file.filename else "png"
    path = f"logirent/logos/{agency_id}/{uuid.uuid4()}.{ext}"
    ct = file.content_type or f"image/{ext}"
    put_object(path, content, ct)

    await db.contract_templates.update_one(
        {"agency_id": agency_id},
        {"$set": {"logo_path": path, "updated_at": datetime.utcnow().isoformat()}},
        upsert=True,
    )
    # Also set agency_id and id if upserting
    existing = await db.contract_templates.find_one({"agency_id": agency_id})
    if existing and not existing.get("id"):
        await db.contract_templates.update_one(
            {"agency_id": agency_id},
            {"$set": {"id": str(uuid.uuid4()), "created_at": datetime.utcnow().isoformat()}}
        )

    return {"logo_path": path}


@router.delete("/admin/contract-template/logo")
async def delete_template_logo(user: dict = Depends(get_agency_admin)):
    agency_id = user.get("agency_id")
    await db.contract_templates.update_one(
        {"agency_id": agency_id},
        {"$set": {"logo_path": None, "updated_at": datetime.utcnow().isoformat()}}
    )
    return {"message": "Logo supprimé"}



# ==================== BOOKING OPTIONS ====================

class BookingOptionInput(PydanticBaseModel):
    name: str
    price_per_day: float
    enabled: bool = True

class BookingOptionsUpdate(PydanticBaseModel):
    options: list[BookingOptionInput]


@router.get("/admin/booking-options")
async def get_booking_options(user: dict = Depends(get_agency_admin)):
    agency_id = user.get("agency_id")
    agency = await db.agencies.find_one({"id": agency_id}, {"_id": 0})
    if not agency:
        raise HTTPException(status_code=404, detail="Agency not found")
    options = agency.get("booking_options", [
        {"name": "GPS", "price_per_day": 10, "enabled": True},
        {"name": "Siège enfant", "price_per_day": 8, "enabled": True},
        {"name": "Conducteur supplémentaire", "price_per_day": 15, "enabled": True},
    ])
    return {"options": options}


@router.put("/admin/booking-options")
async def update_booking_options(data: BookingOptionsUpdate, user: dict = Depends(get_agency_admin)):
    agency_id = user.get("agency_id")
    options = [o.dict() for o in data.options]
    await db.agencies.update_one(
        {"id": agency_id},
        {"$set": {"booking_options": options, "updated_at": datetime.utcnow().isoformat()}}
    )
    return {"options": options, "message": "Options de réservation mises à jour"}


@router.get("/admin/reservations/today")
async def get_today_reservations(user: dict = Depends(get_admin_user)):
    """Get reservations starting today or active today, sorted by start_date"""
    agency_id = user.get('agency_id')
    is_super = user.get('role') == 'super_admin'

    now = datetime.utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = now.replace(hour=23, minute=59, second=59, microsecond=999999)

    base_filter = {} if is_super else {"agency_id": agency_id}
    query = {
        **base_filter,
        "$or": [
            {"start_date": {"$gte": today_start, "$lte": today_end}},
            {"end_date": {"$gte": today_start, "$lte": today_end}},
            {"start_date": {"$lte": today_start}, "end_date": {"$gte": today_end}},
        ]
    }

    reservations = await db.reservations.find(query).sort("start_date", 1).to_list(50)

    user_ids = list(set(r.get('user_id') for r in reservations if r.get('user_id')))
    vehicle_ids = list(set(r.get('vehicle_id') for r in reservations if r.get('vehicle_id')))
    users_list = await db.users.find({"id": {"$in": user_ids}}, {"_id": 0, "id": 1, "name": 1, "email": 1, "phone": 1, "id_photo": 1, "license_photo": 1}).to_list(len(user_ids)) if user_ids else []
    vehicles_list = await db.vehicles.find({"id": {"$in": vehicle_ids}}, {"_id": 0, "id": 1, "brand": 1, "model": 1}).to_list(len(vehicle_ids)) if vehicle_ids else []
    users_map = {u['id']: u for u in users_list}
    vehicles_map = {v['id']: v for v in vehicles_list}

    for res in reservations:
        res['_id'] = str(res['_id'])
        u = users_map.get(res.get('user_id'))
        v = vehicles_map.get(res.get('vehicle_id'))
        res['user_name'] = u['name'] if u else 'Inconnu'
        res['user_email'] = u.get('email', '') if u else ''
        res['user_phone'] = u.get('phone', '') if u else ''
        res['vehicle_name'] = f"{v['brand']} {v['model']}" if v else 'Inconnu'
        res['docs_missing'] = not (u.get('id_photo') and u.get('license_photo')) if u else True

    return {"reservations": reservations, "total": len(reservations)}


@router.get("/admin/reservations/{reservation_id}/check-documents")
async def check_client_documents(reservation_id: str, user: dict = Depends(get_admin_user)):
    """Check if the client for this reservation has all required documents uploaded"""
    reservation = await db.reservations.find_one({"id": reservation_id}, {"_id": 0})
    if not reservation:
        raise HTTPException(status_code=404, detail="Reservation not found")

    client = await db.users.find_one({"id": reservation.get("user_id")}, {"_id": 0, "password_hash": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    missing_docs = []
    if not client.get("id_photo"):
        missing_docs.append("Carte d'identite (recto)")
    if not client.get("id_photo_back"):
        missing_docs.append("Carte d'identite (verso)")
    if not client.get("license_photo"):
        missing_docs.append("Permis de conduire (recto)")
    if not client.get("license_photo_back"):
        missing_docs.append("Permis de conduire (verso)")

    return {
        "client_id": client.get("id"),
        "client_name": client.get("name", ""),
        "documents_complete": len(missing_docs) == 0,
        "missing_documents": missing_docs,
        "has_id": bool(client.get("id_photo")),
        "has_id_back": bool(client.get("id_photo_back")),
        "has_license": bool(client.get("license_photo")),
        "has_license_back": bool(client.get("license_photo_back")),
    }


# ==================== VEHICLE PRICING TIERS ====================

@router.get("/admin/vehicles/{vehicle_id}/pricing")
async def get_vehicle_pricing(vehicle_id: str, user: dict = Depends(get_admin_user)):
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0, "pricing_tiers": 1})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    tiers = vehicle.get("pricing_tiers", [])
    tiers.sort(key=lambda t: t.get("order", 0))
    return {"pricing_tiers": tiers}


@router.put("/admin/vehicles/{vehicle_id}/pricing")
async def update_vehicle_pricing(vehicle_id: str, data: PricingTiersUpdate, user: dict = Depends(get_admin_user)):
    vehicle = await db.vehicles.find_one({"id": vehicle_id})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")

    # Agency admins can only modify their own vehicles
    if user.get('role') != 'super_admin' and vehicle.get('agency_id') != user.get('agency_id'):
        raise HTTPException(status_code=403, detail="Vous ne pouvez modifier que les vehicules de votre agence")

    tiers = []
    for i, t in enumerate(data.pricing_tiers):
        tiers.append({
            "id": t.get("id") or str(uuid.uuid4()),
            "name": t.get("name", ""),
            "kilometers": t.get("kilometers"),
            "price": float(t.get("price", 0)),
            "period": t.get("period", ""),
            "order": t.get("order", i),
            "active": t.get("active", True),
        })

    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$set": {"pricing_tiers": tiers, "updated_at": datetime.utcnow().isoformat()}}
    )
    return {"message": "Tarifs mis a jour", "pricing_tiers": tiers}


# ==================== SEASONAL PRICING ENDPOINTS ====================

@router.get("/admin/vehicles/{vehicle_id}/seasonal-pricing")
async def get_vehicle_seasonal_pricing(vehicle_id: str, user: dict = Depends(get_admin_user)):
    vehicle = await db.vehicles.find_one({"id": vehicle_id})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    pricing = vehicle.get("seasonal_pricing", [])
    return {"seasonal_pricing": pricing}


@router.put("/admin/vehicles/{vehicle_id}/seasonal-pricing")
async def update_vehicle_seasonal_pricing(vehicle_id: str, data: dict, user: dict = Depends(get_admin_user)):
    vehicle = await db.vehicles.find_one({"id": vehicle_id})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")

    if user.get('role') != 'super_admin' and vehicle.get('agency_id') != user.get('agency_id'):
        raise HTTPException(status_code=403, detail="Vous ne pouvez modifier que les vehicules de votre agence")

    seasonal = []
    for s in data.get("seasonal_pricing", []):
        seasonal.append({
            "id": s.get("id") or str(uuid.uuid4()),
            "name": s.get("name", ""),
            "start_date": s.get("start_date", ""),
            "end_date": s.get("end_date", ""),
            "modifier_type": s.get("modifier_type", "percentage"),
            "modifier_value": float(s.get("modifier_value", 0)),
            "active": s.get("active", True),
        })

    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$set": {"seasonal_pricing": seasonal, "updated_at": datetime.utcnow().isoformat()}}
    )
    return {"message": "Tarifs saisonniers mis a jour", "seasonal_pricing": seasonal}


@router.get("/vehicles/{vehicle_id}/seasonal-pricing")
async def get_vehicle_seasonal_pricing_public(vehicle_id: str):
    """Public endpoint: get active seasonal pricing for a vehicle"""
    vehicle = await db.vehicles.find_one({"id": vehicle_id})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    pricing = [s for s in vehicle.get("seasonal_pricing", []) if s.get("active", True)]
    return {"seasonal_pricing": pricing}
